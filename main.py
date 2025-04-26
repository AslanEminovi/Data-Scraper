# main.py

import os
import requests
import openpyxl
import time
import webbrowser
from dotenv import load_dotenv
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from threading import Thread
import re
from tkinter import messagebox, filedialog
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

# Set up the ChromeDriver
CHROMEDRIVER_PATH = r"C:\\Program Files\\chromedriver-win64\\chromedriver.exe"  # Update path if needed
service = Service(CHROMEDRIVER_PATH)

# Import the AISidebar class
from ai_sidebar import AISidebar

# Load the API keys from .env file
load_dotenv()
API_KEY = os.getenv("GOOGLE_MAPS_API_KEY")
RAPIDAPI_KEY = os.getenv("RAPIDAPI_KEY")
RAPIDAPI_HOST = os.getenv("RAPIDAPI_HOST")

if not API_KEY:
    raise ValueError("Google Maps API anahtarı .env dosyasında tanımlı değil!")

# Check RapidAPI credentials (needed for the AISidebar)
if not RAPIDAPI_KEY or not RAPIDAPI_HOST:
    raise ValueError("RapidAPI anahtarı veya host bilgisi .env dosyasında tanımlı değil!")

def format_phone_for_whatsapp(phone):
    """Format phone number for WhatsApp by removing non-numeric characters."""
    if phone and phone != "N/A":
        numbers_only = re.sub(r'\D', '', phone)
        if not numbers_only.startswith('90'):
            numbers_only = '90' + numbers_only
        return numbers_only
    return None

def open_whatsapp(phone):
    """Open WhatsApp with the given phone number."""
    formatted_phone = format_phone_for_whatsapp(phone)
    if formatted_phone:
        whatsapp_url = f"https://api.whatsapp.com/send?phone={formatted_phone}"
        webbrowser.open(whatsapp_url)

def fetch_place_details(place_id):
    """Fetch phone and website information using Place Details API."""
    details_url = (
        "https://maps.googleapis.com/maps/api/place/details/json"
        f"?place_id={place_id}&fields=formatted_phone_number,website"
        f"&key={API_KEY}"
    )
    response = requests.get(details_url)
    details_data = response.json()
    phone = details_data.get("result", {}).get("formatted_phone_number", "N/A")
    website = details_data.get("result", {}).get("website", "N/A")
    return phone, website

def fetch_data():
    """Fetch Google Maps data using a textsearch query."""
    sector = sector_entry.get()
    city = city_entry.get()

    if not sector or not city:
        result_label.config(text="Lütfen sektör ve şehir giriniz.")
        return

    loading_label.set("Veriler çekiliyor, lütfen bekleyin...")

    base_url = (
        "https://maps.googleapis.com/maps/api/place/textsearch/json"
        f"?query={sector}+in+{city}"
        f"&key={API_KEY}&language=tr"
    )

    all_results = []
    next_page_token = None

    while True:
        if next_page_token:
            url = f"{base_url}&pagetoken={next_page_token}"
            time.sleep(2)  # Delay required by Google for next_page_token
        else:
            url = base_url

        try:
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()

            if "results" in data:
                all_results.extend(data["results"])

            next_page_token = data.get("next_page_token")
            if not next_page_token:
                break
        except Exception as e:
            messagebox.showerror("Hata", f"Veri çekerken bir hata oluştu: {str(e)}")
            loading_label.set("")
            return

    if not all_results:
        result_label.config(text="Sonuç bulunamadı.")
        loading_label.set("")
        return

    # Clear Treeview before inserting fresh data
    for row in tree.get_children():
        tree.delete(row)

    for result in all_results:
        name = result.get("name", "N/A")
        address = result.get("formatted_address", "N/A")
        rating = result.get("rating", "N/A")
        place_id = result.get("place_id", None)

        phone, website = fetch_place_details(place_id) if place_id else ("N/A", "N/A")

        tree.insert("", "end", values=(name, sector, rating, address, phone, website))

    result_label.config(text=f"Toplam Veri: {len(all_results)}")
    loading_label.set("")

def start_fetch_data():
    thread = Thread(target=fetch_data)
    thread.start()

def export_to_excel():
    """Export data to Excel with error handling and proper file naming."""
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_filename = f"GoogleMaps_Veriler_{current_time}.xlsx"

    try:
        file_path = filedialog.asksaveasfilename(
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyaları", "*.xlsx")]
        )
        if not file_path:
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Google Maps Verileri"

        headers = ["İşletme Adı", "Sektör/Kategori", "Değerlendirme", "Adres", "Telefon", "Web Sitesi"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)

        for row_idx, item in enumerate(tree.get_children(), 2):
            values = tree.item(item)["values"]
            for col_idx, value in enumerate(values, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 5

        workbook.save(file_path)
        messagebox.showinfo("Başarılı", f"Veriler başarıyla kaydedildi:\n{file_path}")

    except Exception as e:
        messagebox.showerror("Hata", f"Excel dosyası oluşturulurken bir hata oluştu:\n{str(e)}")

def configure_styles():
    """Configure styles for the application."""
    style = ttk.Style()

    # Configure Treeview styles
    style.configure("Treeview", rowheight=30)
    style.configure("Treeview.Heading", rowheight=35)

    return style

def check_whatsapp_all():
    """
    Check each phone in the Treeview to see if it's on WhatsApp.
    If not, display "No WhatsApp" in the phone column.
    Runs in a separate thread so it won't freeze the UI.
    """
    thread = Thread(target=_check_whatsapp_all_impl)
    thread.start()

def _check_whatsapp_all_impl():
    """
    Actual logic to check phone numbers. This is in a separate function
    so we can easily run it in a thread.
    """
    try:
        # Initialize ChromeDriver
        driver = webdriver.Chrome(service=service)

        # 1) Open WhatsApp Web
        driver.get("https://web.whatsapp.com")
        messagebox.showinfo(
            "QR Kod",
            "Lütfen WhatsApp Web'de QR kodunu taratın.\n"
            "Tarama işlemi tamamlandıktan sonra 'Tamam' butonuna basın."
        )

        # Give a few seconds for the user to be fully logged in
        time.sleep(3)

        # 2) Iterate over rows in the Treeview
        for row in tree.get_children():
            values = list(tree.item(row, "values"))  # Convert to list
            phone_index = 4  # "Telefon" is the 5th column (0-based: 4)
            phone_value = values[phone_index]

            if phone_value and phone_value != "N/A":
                # Check with Selenium
                phone_on_whatsapp = check_whatsapp_number(driver, phone_value)
                if not phone_on_whatsapp:
                    # Update the table with "No WhatsApp"
                    values[phone_index] = "No WhatsApp"
                    tree.item(row, values=values)

        driver.quit()

    except Exception as e:
        messagebox.showerror("Hata", f"WhatsApp kontrolünde hata: {str(e)}")

def check_whatsapp_number(driver, phone_number):
    """
    Check if phone_number is on WhatsApp using the open Selenium session.
    Returns True if it is on WhatsApp, False if not.
    """
    formatted = format_phone_for_whatsapp(phone_number)
    if not formatted:
        return False

    # Construct the web.whatsapp.com URL
    url = f"https://web.whatsapp.com/send?phone={formatted}"
    driver.get(url)

    time.sleep(5)  # wait for the page to load or error to show

    # Check if there's an error popup about invalid phone
    try:
        error_box = driver.find_element(By.XPATH, "//div[@data-testid='popup-text']")
        if "phone number shared via url is invalid" in error_box.text.lower():
            return False
    except:
        pass

    # If no error message found, we assume it's on WhatsApp
    return True

# Initialize the main application window
app = ttk.Window(title="Veri Çekme Uygulaması by SynQuad", themename="flatly", size=(1200, 700))
app.resizable(True, True)

# Apply styles
style = configure_styles()

# Create main container
main_container = ttk.Frame(app)
main_container.pack(fill="both", expand=True, padx=20, pady=20)

# Header frame
header_frame = ttk.Frame(main_container)
header_frame.pack(fill="x", pady=(0, 20))

# Input fields frame
input_frame = ttk.Frame(header_frame)
input_frame.pack(fill="x", pady=(0, 10))

# Search fields (Sektör / Şehir)
search_frame = ttk.Frame(input_frame)
search_frame.pack(side="left")

ttk.Label(search_frame, text="İşletme/Sektör:").pack(side="left", padx=5)
sector_entry = ttk.Entry(search_frame, width=30)
sector_entry.pack(side="left", padx=5)

ttk.Label(search_frame, text="Şehir:").pack(side="left", padx=5)
city_entry = ttk.Entry(search_frame, width=30)
city_entry.pack(side="left", padx=5)

# Buttons frame
button_frame = ttk.Frame(input_frame)
button_frame.pack(side="right")

fetch_button = ttk.Button(
    button_frame,
    text="Veri Çek",
    command=start_fetch_data,
    bootstyle="success"
)
fetch_button.pack(side="left", padx=5)

export_button = ttk.Button(
    button_frame,
    text="Excel'e Aktar",
    command=export_to_excel,
    bootstyle="info"
)
export_button.pack(side="left", padx=5)

# NEW: WhatsApp Kontrol button
whatsapp_button = ttk.Button(
    button_frame,
    text="WhatsApp Kontrol",
    command=check_whatsapp_all,
    bootstyle="danger"
)
whatsapp_button.pack(side="left", padx=5)

# AI Sidebar Open button
open_ai_button = ttk.Button(
    button_frame,
    text="Yapay Zeka Aç",  # "Open AI" in Turkish
    command=lambda: AISidebar(app, 50, 50),
    bootstyle="warning"
)
open_ai_button.pack(side="left", padx=5)

# Theme switcher button
theme_button = ttk.Button(
    button_frame,
    text="Koyu Tema",
    command=lambda: style.theme_use("darkly" if style.theme_use() == "flatly" else "flatly"),
    bootstyle="secondary"
)
theme_button.pack(side="left", padx=5)

# Status frame
status_frame = ttk.Frame(main_container)
status_frame.pack(fill="x", pady=(0, 10))

result_label = ttk.Label(status_frame, text="Toplam Veri: 0")
result_label.pack(side="left")

loading_label = ttk.StringVar()
loading_label.set("")
loading_label_widget = ttk.Label(
    status_frame,
    textvariable=loading_label,
    bootstyle="info"
)
loading_label_widget.pack(side="right")

# Table frame (Treeview)
table_frame = ttk.Frame(main_container)
table_frame.pack(fill="both", expand=True)

columns = (
    "İşletme Adı",
    "Sektör/Kategori",
    "Değerlendirme",
    "Adres",
    "Telefon",
    "Web Sitesi"
)
tree = ttk.Treeview(
    table_frame,
    columns=columns,
    show="headings",
    height=20,
    bootstyle="primary"
)

# Set custom column widths
column_widths = {
    "İşletme Adı": 250,
    "Sektör/Kategori": 200,
    "Değerlendirme": 100,
    "Adres": 200,  # Shorter column for the address
    "Telefon": 180,
    "Web Sitesi": 250
}

for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=column_widths.get(col, 150), anchor="center")

tree.pack(fill="both", expand=True)

# Tooltip
tooltip_text = (
    "Telefon numaralarına tıklayarak WhatsApp'ta sohbet başlatabilirsiniz\n"
    "Web sitelerine tıklayarak tarayıcıda açabilirsiniz\n"
    "Adreslere tıklayarak tam adresi görüntüleyebilirsiniz"
)
tooltip = ttk.Label(
    main_container,
    text=tooltip_text,
    bootstyle="info"
)
tooltip.pack(pady=(10, 0))

app.mainloop() 